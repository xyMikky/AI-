"""评论分析共享核心（competitor-analyzer 评论看板 + 静态摘要共用）。

职责
====
1. 统一读入评论数据：`.json`（导出的整份评论）与 `.xlsx`（openpyxl 直读）。
2. 归一化为统一记录结构（兼容中文键 / metadata.items 英文键两种来源）。
3. 数据驱动的关键词抽取（unigram + bigram，停用词过滤）——
   替代旧版写死某产品的 POSITIVE_SIGNALS / CRITICAL_SIGNALS / THEME_PATTERNS。
4. 各维度聚合（星级 / 变体 / 国家 / VP / 月度趋势 / 评论长度）。

设计原则
========
- 不绑定任何品类：关键词、主题均按数据现状产出，可泛化到任意产品。
- 主题（theme）改为**可选**配置：调用方可传入 {主题名: [关键词...]}，
  不传则完全跳过主题，由数据驱动 Top 关键词承担"痛点速览"职责。

归一化记录字段
==============
  stars     int     1-5（无法解析为 0）
  title     str
  body      str     已去除尾部 "Read more Read less"
  text      str     (title + body) 小写，供关键词/主题匹配
  variation str     款式/颜色标签（dict 自动拼为 "Color: Grey" 形态）
  asin      str
  country   str     站点/国家（优先字段，其次从 date_text 解析）
  date      str|None  ISO 日期 YYYY-MM-DD
  month     str|None  YYYY-MM（趋势用）
  vp        bool    是否 Verified Purchase
  length    int     正文字符数
"""

from __future__ import annotations

import collections
import datetime
import json
import re
from pathlib import Path
from typing import Any, Iterable

# ---------------------------------------------------------------------------
# 英文停用词（关键词抽取过滤；评论多为英文，少量他国语言由短词/词频阈值兜底）
# ---------------------------------------------------------------------------
STOPWORDS: set[str] = {
    "the", "and", "for", "are", "but", "not", "you", "all", "any", "can",
    "had", "her", "was", "one", "our", "out", "has", "him", "his", "how",
    "its", "may", "new", "now", "old", "see", "two", "way", "who", "boy",
    "did", "get", "let", "put", "say", "she", "too", "use", "that", "this",
    "with", "have", "from", "they", "will", "would", "there", "their",
    "what", "about", "which", "when", "your", "them", "then", "than",
    "into", "more", "some", "could", "other", "been", "very", "just",
    "also", "after", "only", "over", "such", "most", "even", "make",
    "made", "much", "many", "well", "back", "still", "being", "where",
    "while", "these", "those", "because", "before", "between", "should",
    "does", "doesnt", "dont", "didnt", "isnt", "wasnt", "arent", "werent",
    "cant", "couldnt", "wouldnt", "shouldnt", "wont", "havent", "hasnt",
    "hadnt", "aint", "got", "really", "would", "were", "had", "him", "she",
    "its", "i've", "i'm", "we've", "it", "is", "in", "on", "of", "to", "a",
    "an", "as", "at", "be", "by", "do", "go", "he", "if", "me", "my", "no",
    "or", "so", "up", "us", "we", "am", "read", "less", "more", "im", "ive",
    "id", "ill", "youre", "youve", "theyre", "theyve", "weve", "thats",
    "whats", "heres", "lets", "off", "out", "down", "away", "around",
    "through", "onto", "upon", "again", "once", "here", "both", "each",
    "need", "needs", "needed", "want", "wants", "wanted", "like", "likes",
    "liked", "tried", "try", "trying", "feel", "feels", "felt", "look",
    "looks", "looked", "thought", "think", "used", "using", "uses",
    "actually", "basically", "literally", "definitely", "overall",
    "pretty", "quite", "kind", "sort",
    "product", "amazon", "item", "buy", "bought", "purchase",
    "purchased", "ordered", "order", "review", "reviews", "star", "stars",
    "rating", "one", "thing", "things", "lot", "bit", "way", "time", "times",
}

# 去撇号后按纯字母切词：doesn't→doesnt（被上方停用词拦截），避免花引号截断成 doesn 碎片
WORD_RE = re.compile(r"[a-z]{2,}")
_APOS_RE = re.compile(r"[\u2018\u2019\u02bc']")

# 否定词：不直接丢弃，而是与后面的实词合并成「not work / not good / not worth」，避免语义反转
NEGATORS: set[str] = {
    "not", "no", "never", "none", "cannot", "cant", "couldnt", "wont",
    "wouldnt", "dont", "doesnt", "didnt", "isnt", "wasnt", "arent",
    "werent", "hasnt", "havent", "hadnt", "nor", "neither", "hardly",
    "barely", "without", "aint",
}
_READMORE_RE = re.compile(r"\s*Read more\s*Read less\s*$", re.IGNORECASE)
_COUNTRY_RE = re.compile(r"Reviewed in (?:the )?([A-Za-z][A-Za-z .]+?) on ", re.IGNORECASE)
_ISO_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
_ENDATE_RE = re.compile(r"on\s+([A-Za-z]+)\s+(\d{1,2}),\s+(\d{4})")


def get_field(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return default


def to_int_star(value: Any) -> int:
    try:
        return int(round(float(value)))
    except Exception:
        m = re.search(r"([1-5](?:\.\d+)?)", str(value or ""))
        if not m:
            return 0
        try:
            return int(round(float(m.group(1))))
        except Exception:
            return 0


def _normalize_variation(value: Any) -> str:
    if isinstance(value, dict):
        return " · ".join(
            f"{str(k).capitalize()}: {v}" for k, v in value.items() if v
        )
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, dict):
                label = (
                    item.get("label")
                    or item.get("name")
                    or item.get("key")
                    or item.get("dimension")
                    or ""
                )
                val = (
                    item.get("value")
                    or item.get("selected")
                    or item.get("text")
                    or item.get("option")
                    or ""
                )
                if label and val:
                    parts.append(f"{label}: {val}")
                elif val:
                    parts.append(str(val))
            elif item:
                parts.append(str(item))
        return " · ".join(parts)
    return str(value or "").strip()


def _clean_body(text: str) -> str:
    return _READMORE_RE.sub("", str(text or "")).replace("\n", " ").strip()


def _parse_bool_vp(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("true", "1", "yes", "y", "是", "vp", "verified", "verified purchase")


def parse_review_date(row: dict[str, Any]) -> datetime.date | None:
    """优先中文键 ISO 日期，回退英文 date_text 的 'on Month D, YYYY'。"""
    raw = str(get_field(row, "评论时间", "date", default="")).strip()
    if raw:
        m = _ISO_RE.search(raw)
        if m:
            try:
                return datetime.date(int(m[1]), int(m[2]), int(m[3]))
            except ValueError:
                pass
        try:
            return datetime.datetime.strptime(raw, "%B %d, %Y").date()
        except ValueError:
            pass
    txt = str(get_field(row, "date_text", "评论时间文本", default=""))
    m = _ENDATE_RE.search(txt)
    if m:
        try:
            return datetime.datetime.strptime(f"{m[1]} {m[2]} {m[3]}", "%B %d %Y").date()
        except ValueError:
            return None
    return None


def _extract_country(row: dict[str, Any]) -> str:
    direct = str(
        get_field(row, "所属国家", "country", "国家", "domainCode", default="")
    ).strip()
    if direct:
        return direct
    txt = str(get_field(row, "date_text", default=""))
    m = _COUNTRY_RE.search(txt)
    if m:
        return m.group(1).strip()
    return ""


def normalize_record(row: dict[str, Any]) -> dict[str, Any]:
    title = str(get_field(row, "标题", "title")).strip()
    body = _clean_body(get_field(row, "内容", "body", "text"))
    d = parse_review_date(row)
    return {
        "stars": to_int_star(get_field(row, "星级", "stars", "rating")),
        "title": title,
        "body": body,
        "text": f"{title} {body}".lower(),
        "variation": _normalize_variation(get_field(row, "型号", "variation", "attributes")),
        "asin": str(get_field(row, "ASIN", "asin", default="UNKNOWN")),
        "country": _extract_country(row),
        "date": d.isoformat() if d else None,
        "month": d.strftime("%Y-%m") if d else None,
        "vp": _parse_bool_vp(get_field(row, "VP评论", "verified", "VP", default=False)),
        "length": len(body),
    }


# ---------------------------------------------------------------------------
# 数据读入：.json / .xlsx → 原始行列表
# ---------------------------------------------------------------------------
def _read_json_rows(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        # 兼容 {reviews: {items: [...]}} 或 {items: [...]} 形态
        if "data" in data and isinstance(data["data"], list):
            return data["data"]
        if "items" in data and isinstance(data["items"], list):
            return data["items"]
        reviews = data.get("reviews")
        if isinstance(reviews, dict) and isinstance(reviews.get("items"), list):
            return reviews["items"]
        raise ValueError("JSON 顶层不是数组，且未找到 items 列表")
    if not isinstance(data, list):
        raise ValueError("评论 JSON 顶层应为数组")
    return data


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("读取 .xlsx 需要 openpyxl，请先 pip install openpyxl") from exc
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if raw is None:
            continue
        record = {}
        any_val = False
        for i, key in enumerate(header):
            if not key:
                continue
            val = raw[i] if i < len(raw) else None
            if val not in (None, ""):
                any_val = True
            record[key] = val
        if any_val:
            out.append(record)
    wb.close()
    return out


def resolve_target_asin(
    explicit_asin: str,
    meta_json: dict[str, Any] | None,
    all_variants: bool,
) -> str:
    """解析"该分析哪个 ASIN"。

    默认逻辑（与"只分析 URL 对应竞品"对齐）：
    - 显式 --asin 优先；
    - 否则若未要求 --all-variants 且 metadata.json 提供了 asin，则按该目标 ASIN 过滤；
    - 否则返回 ""（全量，仅当 --all-variants 或无 metadata 时）。
    """
    if explicit_asin:
        return explicit_asin.strip()
    if all_variants:
        return ""
    if isinstance(meta_json, dict):
        a = str(meta_json.get("asin") or "").strip()
        if a:
            return a
    return ""


def global_rating_from_meta(meta_json: dict[str, Any] | None) -> dict[str, Any]:
    """从 metadata.json 提取"全局综合评分"口径（Amazon 页面 global ratings，覆盖全部评论）。

    与"抓取样本口径"严格区分：
    - **全局综合评分**：Amazon 页面对**全部**评论（如 1,126 条）的综合星级与分布，
      不受本次抓取样本量 / 星级配额影响 —— 这才是该 listing 的真实口碑分。
    - **抓取样本**：LinkFox / 用户文件提供的有限样本（每星级上限 100 条，且常按
      正/负配额抓取），低星会被系统性放大，样本均分**不能**当作产品综合分。

    返回字段：
      available              bool   是否拿到全局口径（rating 或 histogram 任一存在）
      rating                 float  页面综合星级（如 4.4）
      rating_count           int    全部评价数（如 1126）
      histogram              list   [{star, percent}] 全局星级百分比分布
      avg                    float  综合均分（优先 rating，缺失时用直方图加权推导）
      avg_from_histogram     float  仅由直方图加权推导的均分（用于校验/兜底）
      low_star_ratio_percent float  全局 1-2★ 占比
      source                 str    口径来源标记
    """
    if not isinstance(meta_json, dict):
        return {"available": False, "source": "amazon_page_global_ratings"}

    rv = meta_json.get("reviews")
    rv = rv if isinstance(rv, dict) else {}
    hist = rv.get("histogram") or []
    rating = meta_json.get("rating")
    rating_count = meta_json.get("rating_count")

    avg_from_hist: float | None = None
    low_pct: float | None = None
    if hist:
        try:
            tot = sum(float(h.get("percent") or 0) for h in hist)
            if tot > 0:
                avg_from_hist = round(
                    sum(float(h["star"]) * float(h.get("percent") or 0) for h in hist)
                    / tot,
                    2,
                )
                low_pct = round(
                    sum(
                        float(h.get("percent") or 0)
                        for h in hist
                        if int(float(h["star"])) <= 2
                    )
                    / tot
                    * 100,
                    1,
                )
        except Exception:
            avg_from_hist = None
            low_pct = None

    return {
        "available": bool(rating is not None or hist),
        "rating": rating,
        "rating_count": rating_count,
        "histogram": hist,
        "avg": rating if rating is not None else avg_from_hist,
        "avg_from_histogram": avg_from_hist,
        "low_star_ratio_percent": low_pct,
        "source": "amazon_page_global_ratings",
    }


def load_reviews(
    input_path: str | Path, asin_filter: str = ""
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """读入并归一化评论。返回 (records, meta)。

    meta 含：source / scope / distinct_asin / distinct_asin_count / raw_count。
    """
    path = Path(input_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"评论文件不存在: {path}")

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        raw_rows = _read_xlsx_rows(path)
    elif suffix == ".json":
        raw_rows = _read_json_rows(path)
    else:
        raise ValueError(f"不支持的评论文件类型: {suffix}（仅支持 .json / .xlsx）")

    records = [normalize_record(r) for r in raw_rows]

    distinct = collections.Counter(r["asin"] for r in records)

    if asin_filter:
        records = [r for r in records if r["asin"] == asin_filter]
        scope = f"single_asin:{asin_filter}"
    else:
        scope = "all_variants"

    meta = {
        "source": str(path),
        "scope": scope,
        "raw_count": len(raw_rows),
        "distinct_asin_count": len(distinct),
        "distinct_asin": dict(distinct.most_common()),
    }
    return records, meta


# ---------------------------------------------------------------------------
# 数据驱动关键词抽取
# ---------------------------------------------------------------------------
def _tokens(text: str) -> list[str]:
    cleaned = _APOS_RE.sub("", str(text or "").lower())
    out: list[str] = []
    neg = False
    span = 0
    for w in WORD_RE.findall(cleaned):
        if w in NEGATORS:          # 标记否定，作用到下一个实词
            neg, span = True, 0
            continue
        if w in STOPWORDS:         # 否定可跨少量虚词（如 not really good）
            if neg:
                span += 1
                if span > 3:
                    neg = False
            continue
        if neg:
            out.append("not " + w)
            neg = False
        else:
            out.append(w)
    return out


def extract_keywords(
    records: Iterable[dict[str, Any]],
    top_n: int = 20,
    min_count: int = 3,
) -> list[dict[str, Any]]:
    """unigram + bigram 词频，返回 [{text, count}]（按词频降序）。

    词组优先：从构成词组的单词计数中扣除词组次数，避免“真正的词组（如
    pressure washer / turbo jet）又被拆成单词重复刷屏”。
    """
    uni: collections.Counter[str] = collections.Counter()
    bi: collections.Counter[str] = collections.Counter()
    for r in records:
        toks = _tokens(r["text"])
        uni.update(toks)
        for a, b in zip(toks, toks[1:]):
            bi[f"{a} {b}"] += 1
    bigrams = {w: c for w, c in bi.items() if c >= 3}
    for phrase, c in bigrams.items():
        for part in phrase.split(" "):
            if part in uni:
                uni[part] -= c
    merged: collections.Counter[str] = collections.Counter()
    for w, c in uni.items():
        if c >= max(min_count, 3):
            merged[w] = c
    for w, c in bigrams.items():
        merged[w] = max(merged.get(w, 0), c)
    return [
        {"text": w, "count": c}
        for w, c in merged.most_common(top_n)
    ]


def keyword_cloud(
    records: list[dict[str, Any]], top_n: int = 18
) -> list[dict[str, Any]]:
    """合并好评(4-5★)/差评(1-2★)关键词为加权词云 [{text,count,sentiment,level}]。"""
    pos = extract_keywords([r for r in records if r["stars"] >= 4], top_n=top_n)
    neg = extract_keywords([r for r in records if 1 <= r["stars"] <= 2], top_n=top_n)
    merged = [(d["text"], d["count"], "pos") for d in pos]
    merged += [(d["text"], d["count"], "neg") for d in neg]
    if not merged:
        return []
    max_c = max(c for _, c, _ in merged)
    cloud = [
        {
            "text": t,
            "count": c,
            "sentiment": s,
            "level": 3 if c / max_c >= 0.66 else (2 if c / max_c >= 0.33 else 1),
        }
        for t, c, s in merged
    ]
    cloud.sort(key=lambda x: -x["count"])
    return cloud


# ---------------------------------------------------------------------------
# 维度聚合
# ---------------------------------------------------------------------------
def _bucket_stats(records: list[dict[str, Any]]) -> dict[str, Any]:
    n = len(records)
    valid = [r["stars"] for r in records if 1 <= r["stars"] <= 5]
    low = sum(1 for r in records if r["stars"] in (1, 2))
    return {
        "count": n,
        "avg": round(sum(valid) / len(valid), 2) if valid else 0.0,
        "low_star_pct": round(low / n * 100, 1) if n else 0.0,
    }


def star_histogram(records: list[dict[str, Any]]) -> dict[str, Any]:
    counts = collections.Counter(r["stars"] for r in records)
    total = len(records)
    hist = [
        {
            "star": s,
            "count": counts.get(s, 0),
            "percent": round(counts.get(s, 0) / total * 100, 1) if total else 0.0,
        }
        for s in (5, 4, 3, 2, 1)
    ]
    return {
        "total": total,
        "avg": _bucket_stats(records)["avg"],
        "low_star_ratio_percent": _bucket_stats(records)["low_star_pct"],
        "star_counts": {str(s): counts.get(s, 0) for s in (5, 4, 3, 2, 1)},
        "histogram": hist,
    }


def group_breakdown(
    records: list[dict[str, Any]], key: str, top: int = 12
) -> list[dict[str, Any]]:
    """按 key（variation / country / asin）分组的均分 + 低星比 + 数量。"""
    groups: dict[str, list[dict[str, Any]]] = collections.defaultdict(list)
    for r in records:
        label = r.get(key) or "（未知）"
        groups[label].append(r)
    out = []
    for label, grp in groups.items():
        st = _bucket_stats(grp)
        out.append({"label": label, **st})
    out.sort(key=lambda x: -x["count"])
    return out[:top]


def vp_breakdown(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    vp = [r for r in records if r["vp"]]
    non = [r for r in records if not r["vp"]]
    return [
        {"label": "Verified Purchase", **_bucket_stats(vp)},
        {"label": "非 VP", **_bucket_stats(non)},
    ]


def monthly_trend(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按月聚合：评论量 + 均分 + 低星比，按月份升序。"""
    months: dict[str, list[dict[str, Any]]] = collections.defaultdict(list)
    for r in records:
        if r["month"]:
            months[r["month"]].append(r)
    out = []
    for m in sorted(months):
        st = _bucket_stats(months[m])
        out.append({"month": m, **st})
    return out


def length_distribution(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """评论正文长度分桶。"""
    buckets = [
        ("0-50", 0, 50),
        ("51-150", 51, 150),
        ("151-300", 151, 300),
        ("301-600", 301, 600),
        ("600+", 601, 10**9),
    ]
    out = []
    for label, lo, hi in buckets:
        grp = [r for r in records if lo <= r["length"] <= hi]
        out.append({"label": label, "count": len(grp), "avg": _bucket_stats(grp)["avg"]})
    return out


def theme_summary(
    records: list[dict[str, Any]], themes: dict[str, list[str]]
) -> dict[str, dict[str, int]]:
    """按可选主题配置统计命中数（全部 + 差评内）。themes 为空则返回 {}。"""
    if not themes:
        return {}
    neg = [r for r in records if r["stars"] in (1, 2)]
    out: dict[str, dict[str, int]] = {}
    for name, pats in themes.items():
        low_pats = [p.lower() for p in pats]
        all_hit = sum(1 for r in records if any(p in r["text"] for p in low_pats))
        neg_hit = sum(1 for r in neg if any(p in r["text"] for p in low_pats))
        out[name] = {"all": all_hit, "negative": neg_hit}
    return out


def real_voice(records: list[dict[str, Any]], per_bucket: int = 3) -> list[dict[str, Any]]:
    """各星级桶择取原话候选（低星优先），含变体信息。"""
    out = []
    for bucket in (1, 2, 3, 5):
        grp = [r for r in records if r["stars"] == bucket]
        for r in grp[:per_bucket]:
            out.append(
                {
                    "stars": r["stars"],
                    "variation": r["variation"],
                    "asin": r["asin"],
                    "text": r["body"][:400],
                }
            )
    return out
