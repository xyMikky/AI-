#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将整个 Excel 工作簿转换为 JSON。

仅依赖 openpyxl（无需 pandas）。支持 .xlsx / .xlsm。
- 单工作表 → 输出对象数组 [ {...}, {...} ]
- 多工作表 → 默认输出 { "工作表名": [ {...} ], ... }（可用 --first-sheet 只取首个）
- 自动把日期/时间单元格转为 ISO 字符串
- 默认输出到与源文件同名的 .json（UTF-8，缩进 2）

用法示例:
    python excel_to_json.py --input data.xlsx
    python excel_to_json.py --input data.xlsx --output out.json --indent 4
    python excel_to_json.py --input data.xlsx --sheet Sheet1 --header-row 2
    python excel_to_json.py --input data.xlsx --orient columns
"""
import argparse
import datetime
import json
import os
import sys


def cell_to_jsonable(value):
    """把单元格值转换为可 JSON 序列化的类型。"""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def make_headers(raw_headers):
    """生成唯一、非空的表头列表。"""
    headers = []
    seen = {}
    for i, h in enumerate(raw_headers):
        name = str(h).strip() if h is not None else ""
        if name == "":
            name = f"col_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    return headers


def sheet_to_records(ws, header_row, orient, number_field=None):
    """将单个工作表转换为 records(对象数组) 或 columns(列字典)。

    number_field 不为空时，给每条数据追加一个从 1 开始的行号字段，
    标注这是第几组数据（records 模式下置于每条记录最前；columns 模式下作为首列）。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [] if orient == "records" else {}

    hr = max(0, header_row - 1) if header_row > 0 else 0
    if hr >= len(rows):
        hr = 0
    headers = make_headers(rows[hr])
    data_rows = rows[hr + 1:]

    records = []
    seq = 0
    for row in data_rows:
        if row is None or all(c is None for c in row):
            continue
        seq += 1
        record = {}
        if number_field:
            record[number_field] = seq
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            record[header] = cell_to_jsonable(value)
        records.append(record)

    if orient == "columns":
        col_headers = ([number_field] if number_field else []) + headers
        columns = {h: [] for h in col_headers}
        for rec in records:
            for h in col_headers:
                columns[h].append(rec.get(h))
        return columns
    return records


def convert(input_path, output_path, sheet, header_row, orient, first_sheet, indent,
            number_field="number"):
    try:
        import openpyxl
    except ImportError:
        print("错误: 缺少依赖 openpyxl。请先运行: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    if not os.path.isfile(input_path):
        print(f"错误: 找不到输入文件 {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    if sheet is not None:
        if sheet not in wb.sheetnames:
            print(f"错误: 工作簿中没有名为 '{sheet}' 的工作表。可用: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
        target_sheets = [sheet]
    else:
        target_sheets = wb.sheetnames

    sheets_data = {}
    for name in target_sheets:
        sheets_data[name] = sheet_to_records(wb[name], header_row, orient, number_field)

    if first_sheet or len(target_sheets) == 1:
        result = sheets_data[target_sheets[0]]
    else:
        result = sheets_data

    if not output_path:
        output_path = os.path.splitext(input_path)[0] + ".json"

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=indent)

    total = sum(len(v) for v in sheets_data.values()) if orient == "records" else None
    summary = {
        "input": input_path,
        "output": output_path,
        "sheets": target_sheets,
        "orient": orient,
        "records_total": total,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return output_path


def main():
    parser = argparse.ArgumentParser(description="将整个 Excel 工作簿转换为 JSON")
    parser.add_argument("--input", "-i", required=True, help="输入 .xlsx/.xlsm 文件路径")
    parser.add_argument("--output", "-o", default=None, help="输出 .json 路径（默认与源文件同名）")
    parser.add_argument("--sheet", "-s", default=None, help="只转换指定名称的工作表（默认全部）")
    parser.add_argument("--header-row", type=int, default=1, help="表头所在行号（1 基，默认 1）")
    parser.add_argument("--orient", choices=["records", "columns"], default="records",
                        help="输出方向: records=对象数组(默认), columns=列字典")
    parser.add_argument("--first-sheet", action="store_true", help="多表时只输出第一个工作表（不包工作表名）")
    parser.add_argument("--indent", type=int, default=2, help="JSON 缩进空格数（默认 2）")
    parser.add_argument("--number-field", default="number",
                        help="行号字段名，标注第几组数据（默认 number）")
    parser.add_argument("--no-number", action="store_true", help="不添加行号字段")
    args = parser.parse_args()

    convert(
        input_path=args.input,
        output_path=args.output,
        sheet=args.sheet,
        header_row=args.header_row,
        orient=args.orient,
        first_sheet=args.first_sheet,
        indent=args.indent,
        number_field=None if args.no_number else args.number_field,
    )


if __name__ == "__main__":
    main()
